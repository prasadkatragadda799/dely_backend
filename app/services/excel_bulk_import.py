"""
Parse .xlsx uploads and create categories, companies, brands, or products (no images).
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.brand import Brand
from app.models.category import Category
from app.models.company import Company
from app.models.division import Division
from app.models.product import Product
from app.utils.slug import generate_slug, make_unique_slug


def _norm_header(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    s = re.sub(r"\s+", "_", s)
    return s


def _collapse_header_underscores(s: str) -> str:
    """Normalize doubled underscores from multi-space headers, e.g. company__name_ -> company_name."""
    if not s:
        return ""
    return re.sub(r"_+", "_", s).strip("_")


def _map_product_column_header(norm: str) -> str:
    """Map normalized Excel headers to canonical product import keys."""
    if not norm:
        return ""
    n = _collapse_header_underscores(norm)
    aliases: Dict[str, str] = {
        "company_name": "company_name",
        "company": "company_name",
        "products": "products_line",
        "product": "products_line",
        "product_name": "name",
        "name": "name",
        "item": "name",
        "item_name": "name",
        "title": "name",
        "vireyant": "variant_suffix",
        "variant": "variant_suffix",
        "flavour": "variant_suffix",
        "flavor": "variant_suffix",
        "mrp": "mrp",
        "rsp": "selling_price",
        "selling_price": "selling_price",
        "sale_price": "selling_price",
        "sp": "selling_price",
        "price": "selling_price",
        "size": "unit",
        "pack": "unit",
        "pack_size": "unit",
        "uom": "unit",
        "unit": "unit",
        "qty": "stock_quantity",
        "quantity": "stock_quantity",
        "stock": "stock_quantity",
        "stock_quantity": "stock_quantity",
        "exp_date": "expiry_date",
        "expiry": "expiry_date",
        "expiry_date": "expiry_date",
        "best_before": "expiry_date",
        "mfg_date": "manufacturing_date",
        "muf_date": "manufacturing_date",
        "manufacturing_date": "manufacturing_date",
    }
    if n in aliases:
        return aliases[n]
    if "exp" in n and "date" in n:
        return "expiry_date"
    if ("muf" in n or "mfg" in n) and "date" in n:
        return "manufacturing_date"
    return n


def _map_generic_import_header(norm: str) -> str:
    """Map headers for companies / categories / brands imports."""
    if not norm:
        return ""
    n = _collapse_header_underscores(norm)
    aliases: Dict[str, str] = {
        "name": "name",
        "description": "description",
        "company_name": "company_name",
        "company": "company_name",
        "category_name": "category_name",
        "category": "category_name",
        "parent_name": "parent_name",
        "division_slug": "division_slug",
        "display_order": "display_order",
        "is_active": "is_active",
        "icon": "icon",
        "color": "color",
        "meta_title": "meta_title",
        "meta_description": "meta_description",
    }
    return aliases.get(n, n)


def _score_product_header_row(norms: List[str]) -> int:
    mapped = {_map_product_column_header(n) for n in norms if n}
    mapped.discard("")
    score = 0
    if "mrp" in mapped:
        score += 5
    if "name" in mapped or "products_line" in mapped:
        score += 3
    if "company_name" in mapped:
        score += 2
    if "unit" in mapped:
        score += 2
    if "variant_suffix" in mapped:
        score += 1
    if "stock_quantity" in mapped:
        score += 1
    if "expiry_date" in mapped:
        score += 1
    if "selling_price" in mapped:
        score += 1
    return score


def _score_non_product_header_row(norms: List[str], import_entity: str) -> int:
    mapped = {_map_generic_import_header(n) for n in norms if n}
    mapped.discard("")
    score = 0
    if "name" in mapped:
        score += 5
    if import_entity == "brands":
        if "company_name" in mapped:
            score += 3
        if "category_name" in mapped:
            score += 3
    if "description" in mapped:
        score += 1
    if "parent_name" in mapped:
        score += 2
    return score


def _detect_header_row_index(rows: List[Any], import_entity: Optional[str]) -> int:
    """Pick header row when sheet has title rows above column labels."""
    if import_entity is None or not rows:
        return 0
    limit = min(20, len(rows))
    best_i = 0
    best_s = -1
    for i in range(limit):
        norms = [_norm_header(c) for c in rows[i]]
        if import_entity == "products":
            s = _score_product_header_row(norms)
        else:
            s = _score_non_product_header_row(norms, import_entity)
        if s > best_s:
            best_s = s
            best_i = i
    if best_s <= 0:
        return 0
    return best_i


def _product_import_display_name(row: Dict[str, Any]) -> Optional[str]:
    """Single name column, or combine products_line + variant_suffix (common in retail listings)."""
    n = _cell_str(row.get("name"))
    if n:
        return n
    parts: List[str] = []
    pl = _cell_str(row.get("products_line"))
    vs = _cell_str(row.get("variant_suffix"))
    if pl:
        parts.append(pl)
    if vs:
        parts.append(vs)
    if parts:
        return " ".join(parts)
    return None


def _cell_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, (datetime, date)):
        return val.isoformat()[:10] if hasattr(val, "isoformat") else str(val)
    s = str(val).strip()
    return s if s else None


def _cell_decimal(val: Any) -> Optional[Decimal]:
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return val
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _cell_int(val: Any, default: Optional[int] = None) -> Optional[int]:
    if val is None or val == "":
        return default
    try:
        if isinstance(val, float):
            return int(val)
        return int(val)
    except (ValueError, TypeError):
        return default


def _parse_bool(val: Any, default: bool = True) -> bool:
    if val is None or val == "":
        return default
    s = str(val).strip().lower()
    if s in ("true", "1", "yes", "y"):
        return True
    if s in ("false", "0", "no", "n"):
        return False
    return default


def parse_excel_rows(
    file_bytes: bytes,
    *,
    import_entity: Optional[str] = None,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    from openpyxl import load_workbook

    # read_only=False so we can scan for header row then read body (files capped at ~5MB by API).
    wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    try:
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []

        header_idx = _detect_header_row_index(all_rows, import_entity)
        raw_header = all_rows[header_idx]
        norms = [_norm_header(h) for h in raw_header]
        while norms and not norms[-1]:
            norms.pop()

        if import_entity == "products":
            headers = [_map_product_column_header(n) for n in norms]
        elif import_entity in ("companies", "categories", "brands"):
            headers = [_map_generic_import_header(n) for n in norms]
        else:
            headers = norms

        out: List[Dict[str, Any]] = []
        for row in all_rows[header_idx + 1 :]:
            if not row:
                continue
            cells = [row[i] if i < len(row) else None for i in range(len(headers))]
            if all(v is None or (isinstance(v, str) and not str(v).strip()) for v in cells):
                continue
            d: Dict[str, Any] = {}
            for i, key in enumerate(headers):
                if not key:
                    continue
                d[key] = row[i] if i < len(row) else None
            out.append(d)
        return headers, out
    finally:
        wb.close()


def build_template_workbook(sheet_name: str, headers: List[str], example_row: List[Any]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    ws.append(example_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_company_by_name(db: Session, name: str) -> Optional[Company]:
    n = name.strip()
    if not n:
        return None
    return (
        db.query(Company)
        .filter(func.lower(Company.name) == func.lower(n))
        .first()
    )


def _find_category_by_name(db: Session, name: str) -> Optional[Category]:
    n = name.strip()
    if not n:
        return None
    return (
        db.query(Category)
        .filter(func.lower(Category.name) == func.lower(n))
        .first()
    )


def _find_brand(db: Session, name: str, company_id: Optional[str] = None) -> Optional[Brand]:
    n = name.strip()
    if not n:
        return None
    q = db.query(Brand).filter(func.lower(Brand.name) == func.lower(n))
    if company_id:
        q = q.filter(Brand.company_id == company_id)
    else:
        q = q.filter(Brand.company_id.is_(None))
    return q.first()


def _find_brand_any_company(db: Session, name: str) -> Optional[Brand]:
    n = name.strip()
    if not n:
        return None
    return db.query(Brand).filter(func.lower(Brand.name) == func.lower(n)).first()


def _get_by_id_variants(db: Session, model, raw_id: str):
    if not raw_id:
        return None
    s = str(raw_id).strip()
    for candidate in (s, s.replace("-", "")):
        obj = db.query(model).filter(model.id == candidate).first()
        if obj:
            return obj
    return None


def _find_division_by_slug(db: Session, slug: Optional[str]) -> Optional[str]:
    if not slug or not str(slug).strip():
        return None
    s = str(slug).strip().lower()
    div = db.query(Division).filter(func.lower(Division.slug) == s, Division.is_active == True).first()
    return str(div.id) if div else None


def import_companies(db: Session, rows: List[Dict[str, Any]]) -> Tuple[int, List[Dict[str, Any]]]:
    created = 0
    errors: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        name = _cell_str(row.get("name"))
        if not name:
            errors.append({"row": idx, "error": "name is required"})
            continue
        if _find_company_by_name(db, name):
            errors.append({"row": idx, "error": f"Company already exists: {name}"})
            continue
        desc = _cell_str(row.get("description"))
        company = Company(name=name, description=desc, logo_url=None)
        db.add(company)
        try:
            db.commit()
            db.refresh(company)
            created += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": idx, "error": str(e)})
    return created, errors


def import_categories(db: Session, rows: List[Dict[str, Any]]) -> Tuple[int, List[Dict[str, Any]]]:
    created = 0
    errors: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        name = _cell_str(row.get("name"))
        if not name:
            errors.append({"row": idx, "error": "name is required"})
            continue
        parent_id = None
        pn = _cell_str(row.get("parent_name"))
        if pn:
            parent = _find_category_by_name(db, pn)
            if not parent:
                errors.append({"row": idx, "error": f"Parent category not found: {pn}"})
                continue
            parent_id = parent.id

        existing = (
            db.query(Category)
            .filter(func.lower(Category.name) == name.lower(), Category.parent_id == parent_id)
            .first()
        )
        if existing:
            errors.append({"row": idx, "error": f"Category already exists at this level: {name}"})
            continue

        division_id = _find_division_by_slug(db, _cell_str(row.get("division_slug")))
        slug_base = generate_slug(name)
        existing_slugs = [c.slug for c in db.query(Category.slug).all()]
        slug = make_unique_slug(slug_base, existing_slugs)
        display_order = _cell_int(row.get("display_order"), 0) or 0
        is_active = _parse_bool(row.get("is_active"), True)
        description = _cell_str(row.get("description"))
        icon = _cell_str(row.get("icon"))
        if icon and len(icon) > 10:
            icon = icon[:10]
        raw_color = _cell_str(row.get("color"))
        color = raw_color if raw_color and re.match(r"^#[0-9A-Fa-f]{6}$", raw_color.strip()) else None
        meta_title = _cell_str(row.get("meta_title"))
        meta_description = _cell_str(row.get("meta_description"))

        cat = Category(
            name=name,
            slug=slug,
            description=description,
            parent_id=parent_id,
            division_id=division_id,
            icon=icon,
            color=color,
            display_order=display_order,
            is_active=is_active,
            meta_title=meta_title,
            meta_description=meta_description,
        )
        db.add(cat)
        try:
            db.commit()
            db.refresh(cat)
            created += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": idx, "error": str(e)})
    return created, errors


def import_brands(db: Session, rows: List[Dict[str, Any]]) -> Tuple[int, List[Dict[str, Any]]]:
    created = 0
    errors: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        name = _cell_str(row.get("name"))
        if not name:
            errors.append({"row": idx, "error": "name is required"})
            continue
        cn = _cell_str(row.get("company_name"))
        if not cn:
            errors.append({"row": idx, "error": "company_name is required"})
            continue
        comp = _find_company_by_name(db, cn)
        if not comp:
            errors.append({"row": idx, "error": f"Company not found: {cn}"})
            continue
        company_id_str = str(comp.id)

        catn = _cell_str(row.get("category_name"))
        if not catn:
            errors.append({"row": idx, "error": "category_name is required"})
            continue
        cat = _find_category_by_name(db, catn)
        if not cat:
            errors.append({"row": idx, "error": f"Category not found: {catn}"})
            continue
        category_id_str = str(cat.id)

        q = db.query(Brand).filter(func.lower(Brand.name) == name.lower())
        if company_id_str:
            q = q.filter(Brand.company_id == company_id_str)
        else:
            q = q.filter(Brand.company_id.is_(None))
        if q.first():
            errors.append({"row": idx, "error": f"Brand already exists (same name/company scope): {name}"})
            continue

        brand = Brand(name=name, company_id=company_id_str, category_id=category_id_str, logo_url=None)
        db.add(brand)
        try:
            db.commit()
            db.refresh(brand)
            created += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": idx, "error": str(e)})
    return created, errors


def _parse_expiry(val: Any) -> Optional[date]:
    s = _cell_str(val)
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def import_products(db: Session, rows: List[Dict[str, Any]], created_by_admin_id: Optional[str]) -> Tuple[int, List[Dict[str, Any]]]:
    created = 0
    errors: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        name = _product_import_display_name(row)
        if not name:
            errors.append({"row": idx, "error": "name is required"})
            continue
        mrp = _cell_decimal(row.get("mrp"))
        sp = _cell_decimal(row.get("selling_price"))
        if mrp is None or mrp <= 0:
            errors.append({"row": idx, "error": "mrp must be a positive number"})
            continue
        if sp is None or sp <= 0:
            sp = mrp
        if sp > mrp:
            errors.append({"row": idx, "error": "selling_price cannot exceed mrp"})
            continue
        unit = _cell_str(row.get("unit"))
        if not unit:
            errors.append({"row": idx, "error": "unit is required (e.g. piece, kg)"})
            continue

        category_id_str: Optional[str] = None
        cid = _cell_str(row.get("category_id"))
        if cid:
            cat = _get_by_id_variants(db, Category, cid)
            if not cat:
                errors.append({"row": idx, "error": f"category_id not found: {cid}"})
                continue
            category_id_str = str(cat.id)
        else:
            catn = _cell_str(row.get("category_name"))
            if catn:
                cat = _find_category_by_name(db, catn)
                if not cat:
                    errors.append({"row": idx, "error": f"category not found: {catn}"})
                    continue
                category_id_str = str(cat.id)

        company_id_str: Optional[str] = None
        coid = _cell_str(row.get("company_id"))
        if coid:
            comp = _get_by_id_variants(db, Company, coid)
            if not comp:
                errors.append({"row": idx, "error": f"company_id not found: {coid}"})
                continue
            company_id_str = str(comp.id)
        else:
            cn = _cell_str(row.get("company_name"))
            if cn:
                comp = _find_company_by_name(db, cn)
                if not comp:
                    errors.append({"row": idx, "error": f"company not found: {cn}"})
                    continue
                company_id_str = str(comp.id)

        brand_id_str: Optional[str] = None
        bid = _cell_str(row.get("brand_id"))
        if bid:
            br = _get_by_id_variants(db, Brand, bid)
            if not br:
                errors.append({"row": idx, "error": f"brand_id not found: {bid}"})
                continue
            brand_id_str = str(br.id)
        else:
            bn = _cell_str(row.get("brand_name"))
            if bn:
                br = _find_brand(db, bn, company_id_str) if company_id_str else None
                if not br:
                    br = _find_brand_any_company(db, bn)
                if not br:
                    errors.append({"row": idx, "error": f"brand not found: {bn}"})
                    continue
                brand_id_str = str(br.id)

        division_id_str = _find_division_by_slug(db, _cell_str(row.get("division_slug")))
        stock = _cell_int(row.get("stock_quantity"), 0) or 0
        min_o = _cell_int(row.get("min_order_quantity"), 1) or 1
        if min_o < 1:
            min_o = 1
        pieces = _cell_int(row.get("pieces_per_set"), 1) or 1
        if pieces < 1:
            pieces = 1
        comm = _cell_decimal(row.get("commission_cost")) or Decimal("0")
        if comm < 0:
            errors.append({"row": idx, "error": "commission_cost cannot be negative"})
            continue
        description = _cell_str(row.get("description"))
        hsn = _cell_str(row.get("hsn_code"))
        is_featured = _parse_bool(row.get("is_featured"), False)
        is_available = _parse_bool(row.get("is_available"), True)
        expiry = _parse_expiry(row.get("expiry_date"))
        meta_title = _cell_str(row.get("meta_title"))
        meta_description = _cell_str(row.get("meta_description"))

        product_slug = generate_slug(name)
        existing_slugs = [p.slug for p in db.query(Product.slug).all()]
        product_slug = make_unique_slug(product_slug, existing_slugs)

        product = Product(
            name=name,
            slug=product_slug,
            description=description,
            brand_id=brand_id_str,
            company_id=company_id_str,
            category_id=category_id_str,
            division_id=division_id_str,
            mrp=mrp,
            selling_price=sp,
            commission_cost=comm,
            stock_quantity=stock,
            min_order_quantity=min_o,
            unit=unit,
            pieces_per_set=pieces,
            specifications=None,
            is_featured=is_featured,
            is_available=is_available,
            expiry_date=expiry,
            meta_title=meta_title,
            meta_description=meta_description,
            hsn_code=hsn,
            created_by=created_by_admin_id,
        )
        db.add(product)
        try:
            db.commit()
            db.refresh(product)
            created += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": idx, "error": str(e)})
    return created, errors
